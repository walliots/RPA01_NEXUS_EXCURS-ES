import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from datetime import datetime
from src.utils.logger import setup_logger
from src.utils.config_loader import load_config

logger = setup_logger("OutputRobot")

# Paleta de cores por ônibus (cicla se houver mais de 8 ônibus)
CORES_ONIBUS = [
    "D6E4F0", "D5F5E3", "FDEBD0", "F9EBEA", "EAF2F8",
    "F4ECF7", "FEF9E7", "E8F8F5",
]
COR_CABECALHO = "2C3E50"
COR_TEXTO_CABECALHO = "FFFFFF"
COR_AVISO = "F0B27A"


class OutputRobot:
    """
    Gera a planilha de saída formatada com a distribuição nos ônibus,
    uma aba resumo e validação dos dados finais.
    """

    def __init__(self):
        self.config = load_config()
        self.cols = self.config["colunas_entrada"]
        self.prefixo = self.config["onibus"]["prefixo_nome"]
        self.capacidade = self.config["onibus"]["capacidade_maxima"]

    def executar(self, df: pd.DataFrame, caminho_saida: str) -> str:
        logger.info("[OUTPUT] Iniciando geração do arquivo de saída...")
        os.makedirs(os.path.dirname(caminho_saida), exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)

        self._criar_aba_resumo(wb, df)
        self._criar_abas_onibus(wb, df)
        self._criar_aba_completa(wb, df)

        wb.save(caminho_saida)
        logger.info(f"[OUTPUT] Arquivo gerado com sucesso: {caminho_saida}")
        return caminho_saida

    # ─────────────────────────── ABA RESUMO ────────────────────────────

    def _criar_aba_resumo(self, wb: Workbook, df: pd.DataFrame):
        ws = wb.create_sheet("📊 Resumo")
        ws.sheet_view.showGridLines = False

        # Título
        ws.merge_cells("A1:E1")
        ws["A1"] = "🚌 DISTRIBUIÇÃO DE EXCURSÃO — RESUMO"
        ws["A1"].font = Font(bold=True, size=14, color=COR_TEXTO_CABECALHO)
        ws["A1"].fill = PatternFill("solid", fgColor=COR_CABECALHO)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A2"].font = Font(italic=True, size=9, color="888888")
        ws.merge_cells("A2:E2")
        ws["A2"].alignment = Alignment(horizontal="center")

        # Totais gerais
        total_pax = len(df)
        total_onibus = df["__onibus_num"].max()
        total_grupos = df["__grupo_id"].nunique()

        dados_gerais = [
            ("Total de Passageiros", total_pax),
            ("Total de Ônibus", int(total_onibus)),
            ("Grupos de Afinidade", total_grupos),
            ("Capacidade por Ônibus", self.capacidade),
        ]

        ws["A4"] = "INDICADORES GERAIS"
        ws["A4"].font = Font(bold=True, size=10, color=COR_TEXTO_CABECALHO)
        ws["A4"].fill = PatternFill("solid", fgColor="5D6D7E")
        ws.merge_cells("A4:B4")

        for i, (label, valor) in enumerate(dados_gerais, start=5):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = valor
            ws[f"A{i}"].font = Font(bold=True, size=10)
            ws[f"B{i}"].alignment = Alignment(horizontal="center")

        # Tabela por ônibus
        ws["A10"] = "DETALHAMENTO POR ÔNIBUS"
        ws["A10"].font = Font(bold=True, size=10, color=COR_TEXTO_CABECALHO)
        ws["A10"].fill = PatternFill("solid", fgColor="5D6D7E")
        ws.merge_cells("A10:E10")

        cabecalhos = ["Ônibus", "Rota", "Passageiros", "Capacidade", "Ocupação (%)", "Status"]
        for col, cab in enumerate(cabecalhos, start=1):
            cell = ws.cell(row=11, column=col, value=cab)
            cell.font = Font(bold=True, color=COR_TEXTO_CABECALHO, size=9)
            cell.fill = PatternFill("solid", fgColor=COR_CABECALHO)
            cell.alignment = Alignment(horizontal="center")

        linha = 12
        for num_onibus in sorted(df["__onibus_num"].unique()):
            df_bus = df[df["__onibus_num"] == num_onibus]
            qtd = len(df_bus)
            tipo = df_bus["__onibus_tipo"].iloc[0] if "__onibus_tipo" in df_bus.columns else ""
            status = "✅ OK" if qtd <= self.capacidade else "⚠️ Excedido"
            cor_linha = CORES_ONIBUS[(int(num_onibus) - 1) % len(CORES_ONIBUS)]

            valores = [f"{self.prefixo} {int(num_onibus):02d}", tipo, qtd, self.capacidade, f"=D{linha}/E{linha}", status]
            for col, val in enumerate(valores, start=1):
                cell = ws.cell(row=linha, column=col, value=val)
                cell.fill = PatternFill("solid", fgColor=cor_linha if status != "⚠️ Excedido" else COR_AVISO)
                cell.alignment = Alignment(horizontal="center")
                if col == 5:
                    cell.number_format = "0.0%"
            linha += 1

        for col, w in zip(["A", "B", "C", "D", "E", "F"], [14, 14, 14, 14, 14, 12]):
            ws.column_dimensions[col].width = w

    # ─────────────────────────── ABA POR ÔNIBUS ────────────────────────

    def _criar_abas_onibus(self, wb: Workbook, df: pd.DataFrame):
        for num_onibus in sorted(df["__onibus_num"].unique()):
            nome_aba = f"🚌 Ônibus {int(num_onibus):02d}"
            ws = wb.create_sheet(nome_aba)
            ws.sheet_view.showGridLines = False

            df_onibus = df[df["__onibus_num"] == num_onibus].copy()
            tipo = df_onibus["__onibus_tipo"].iloc[0] if "__onibus_tipo" in df_onibus.columns else ""
            cor = CORES_ONIBUS[(int(num_onibus) - 1) % len(CORES_ONIBUS)]

            # Título da aba
            ws.merge_cells("A1:F1")
            ws["A1"] = f"🚌 {self.prefixo} {int(num_onibus):02d}  [{tipo}]  —  {len(df_onibus)} passageiro(s)"
            ws["A1"].font = Font(bold=True, size=12, color=COR_TEXTO_CABECALHO)
            ws["A1"].fill = PatternFill("solid", fgColor=COR_CABECALHO)
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28

            # Cabeçalho da tabela
            colunas = ["#", "Nome Completo", "CPF", "Ponto de Embarque", "WhatsApp", "Grupo de Afinidade"]
            larguras = [5, 35, 16, 22, 15, 18]

            for col_idx, (cab, larg) in enumerate(zip(colunas, larguras), start=1):
                cell = ws.cell(row=2, column=col_idx, value=cab)
                cell.font = Font(bold=True, color=COR_TEXTO_CABECALHO, size=9)
                cell.fill = PatternFill("solid", fgColor="5D6D7E")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions[get_column_letter(col_idx)].width = larg

            ws.row_dimensions[2].height = 20

            # Dados ordenados por grupo de afinidade
            df_onibus_sorted = df_onibus.sort_values("__grupo_id")
            grupo_anterior = None

            for linha_idx, (_, row) in enumerate(df_onibus_sorted.iterrows(), start=3):
                grupo_atual = row["__grupo_id"]
                # Alterna cor ao mudar de grupo para facilitar leitura visual
                cor_linha = cor if grupo_atual != grupo_anterior and grupo_anterior is not None else cor
                grupo_anterior = grupo_atual

                valores = [
                    linha_idx - 2,
                    row[self.cols["nome"]],
                    row[self.cols["cpf"]],
                    row[self.cols["ponto_embarque"]],
                    row[self.cols["whatsapp"]],
                    f"Grupo {int(grupo_atual):03d}",
                ]
                for col_idx, val in enumerate(valores, start=1):
                    cell = ws.cell(row=linha_idx, column=col_idx, value=val)
                    cell.fill = PatternFill("solid", fgColor=cor_linha)
                    cell.alignment = Alignment(horizontal="left" if col_idx == 2 else "center")
                    cell.font = Font(size=9)

            ws.freeze_panes = "A3"

    # ─────────────────────────── ABA COMPLETA ──────────────────────────

    def _criar_aba_completa(self, wb: Workbook, df: pd.DataFrame):
        ws = wb.create_sheet("📋 Lista Completa")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:G1")
        ws["A1"] = "LISTA COMPLETA DE PASSAGEIROS"
        ws["A1"].font = Font(bold=True, size=12, color=COR_TEXTO_CABECALHO)
        ws["A1"].fill = PatternFill("solid", fgColor=COR_CABECALHO)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        colunas = ["Ônibus", "Rota", "Nome Completo", "CPF", "Ponto de Embarque", "Email", "WhatsApp", "Grupo"]
        larguras = [14, 12, 35, 16, 22, 30, 14, 12]

        for col_idx, (cab, larg) in enumerate(zip(colunas, larguras), start=1):
            cell = ws.cell(row=2, column=col_idx, value=cab)
            cell.font = Font(bold=True, color=COR_TEXTO_CABECALHO, size=9)
            cell.fill = PatternFill("solid", fgColor="5D6D7E")
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = larg

        df_sorted = df.sort_values(["__onibus_num", "__grupo_id"]).reset_index(drop=True)

        for linha_idx, (_, row) in enumerate(df_sorted.iterrows(), start=3):
            num = int(row["__onibus_num"])
            tipo = row["__onibus_tipo"] if "__onibus_tipo" in row else ""
            cor = CORES_ONIBUS[(num - 1) % len(CORES_ONIBUS)]
            valores = [
                f"{self.prefixo} {num:02d}",
                tipo,
                row[self.cols["nome"]],
                row[self.cols["cpf"]],
                row[self.cols["ponto_embarque"]],
                row[self.cols["email"]],
                row[self.cols["whatsapp"]],
                f"Grupo {int(row['__grupo_id']):03d}",
            ]
            for col_idx, val in enumerate(valores, start=1):
                cell = ws.cell(row=linha_idx, column=col_idx, value=val)
                cell.fill = PatternFill("solid", fgColor=cor)
                cell.font = Font(size=9)
                cell.alignment = Alignment(horizontal="left" if col_idx in [3, 6] else "center")

        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:H{len(df_sorted) + 2}"
